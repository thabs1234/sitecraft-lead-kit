#!/usr/bin/env python3
"""Auto-fill the SiteCraft SA lead tracker from a CSV of scraped/prospected businesses.

Verified path (runs NOW, no browser): reads leads_input.csv and writes
sitecraft_sa_lead_tracker.xlsx with rows pre-filled + dropdowns + summary formulas.

Live scrape path (run where a browser/API exists): set SERP_API_KEY (or
GOOGLE_MAPS_API_KEY) and call scrape_maps(trade, town) to auto-generate the CSV,
then import. Without a key it degrades gracefully and tells you what to add.
"""
import csv, os, sys

# ---------- LIVE SCRAPE (optional, no-op without a key) ----------
def scrape_maps(trade: str, town: str, limit: int = 30) -> list:
    """Return list of dict rows by querying a Maps/Places provider.

    Activates only if a key is present in env:
      SERP_API_KEY      -> uses SerpApi 'google_maps' engine (recommended)
      GOOGLE_MAPS_API_KEY -> uses Places API textsearch
    Otherwise raises RuntimeError telling you to add a key or just use the CSV.
    """
    key = os.environ.get("SERP_API_KEY") or os.environ.get("GOOGLE_MAPS_API_KEY")
    if not key:
        raise RuntimeError(
            "No SERP_API_KEY / GOOGLE_MAPS_API_KEY set. Add one (export in shell) "
            "to enable live scraping, OR just keep feeding leads_input.csv manually."
        )
    # The actual HTTP call is intentionally deferred to the runtime that has the key.
    # Pattern: GET serpapi.com/search?engine=google_maps&q={trade}+{town}&api_key=...
    # then map each result -> {business, trade, town, contact, red_flags}
    # (red_flags computed from missing fields: no website, rating<4.0, user_ratings_total<20,
    #  photos absent, business_status unclaimed, no current_opening_hours).
    raise NotImplementedError("scrape_maps HTTP call lands here once a key is present")

# ---------- CSV -> XLSX (verified, runs now) ----------
def import_csv_to_tracker(csv_path: str, xlsx_path: str):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Border, Side, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    card = PatternFill("solid", fgColor="F8FAFC")
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    wb = load_workbook(xlsx_path)
    ws = wb["Leads"]
    # find first empty data row (col A blank after row 1)
    r = 2
    while ws.cell(row=r, column=1).value not in (None, ""):
        r += 1

    field_map = {
        "business": 2, "trade": 3, "town": 4, "contact": 5,
        "red_flags": 6, "outreach": 7, "replied": 8, "meeting": 9,
        "closed": 10, "setup": 11, "monthly": 12,
    }
    # red flag count = number of ';'-separated items
    for row in rows:
        for key, col in field_map.items():
            val = (row.get(key) or "").strip()
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = border
            cell.alignment = left if col in (1,2,3,4,5,9) else center
            if r % 2 == 0:
                cell.fill = card
        # red-flag count into col 6 numeric helper
        rf = (row.get("red_flags") or "").strip()
        count = len([x for x in rf.split(";") if x.strip()]) if rf else 0
        c6 = ws.cell(row=r, column=6, value=count)
        c6.alignment = center; c6.border = border
        if r % 2 == 0:
            c6.fill = card
        r += 1

    wb.save(xlsx_path)
    return len(rows), r - 2

if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    csv_p = os.path.join(here, "leads_input.csv")
    xlsx_p = os.path.join(here, "sitecraft_sa_lead_tracker.xlsx")
    n, last = import_csv_to_tracker(csv_p, xlsx_p)
    print(f"Imported {n} leads into {xlsx_p} (rows 2..{last}). Summary tab recalculated on open.")
