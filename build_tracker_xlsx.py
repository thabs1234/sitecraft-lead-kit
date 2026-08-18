#!/usr/bin/env python3
"""Build sitecraft_sa_lead_tracker.xlsx — a working lead tracker with formulas + dropdowns."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

ACCENT = "16A34A"; CARD = "F8FAFC"; BORDER = "E2E8F0"; GOLD = "F59E0B"
hdr_fill = PatternFill("solid", fgColor=ACCENT)
hdr_font = Font(bold=True, color="FFFFFF", size=10)
card_fill = PatternFill("solid", fgColor=CARD)
thin = Side(style="thin", color=BORDER)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ---------- Sheet 1: Leads ----------
ws = wb.active
ws.title = "Leads"
cols = ["Date","Business","Trade","Town","Contact (WA/phone)","Red flags (count)","Outreach sent?","Replied?","Meeting/Follow-up","Closed?","Setup value (R)","Monthly value (R)"]
ncol = len(cols)
ws.append(cols)
for c in range(1, ncol+1):
    cell = ws.cell(row=1, column=c)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center; cell.border = border

# pre-fill 40 rows with formulas
N = 40
for r in range(2, N+2):
    # red-flag helper column formula lives in col 6 (user types count)
    ws.cell(row=r, column=6).alignment = center
    for c in range(1, ncol+1):
        cell = ws.cell(row=r, column=c)
        cell.border = border
        if c in (1,2,3,4,5,9):
            cell.alignment = left
        else:
            cell.alignment = center
        if r % 2 == 0:
            cell.fill = card_fill

# column letters
Col = {name:get_column_letter(i+1) for i,name in enumerate(cols)}
SETUP = Col["Setup value (R)"]; MON = Col["Monthly value (R)"]; CLOSED = Col["Closed?"]

# dropdown validations
def add_dv(formula_list, col, prompt):
    dv = DataValidation(type="list", formula1=f'"{formula_list}"', allow_blank=True)
    dv.prompt = prompt; dv.promptTitle = col
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{N+1}")

add_dv("Yes,No", Col["Outreach sent?"], "Did you send the WhatsApp/opener?")
add_dv("Yes,No", Col["Replied?"], "Did they reply?")
add_dv("Yes,No,Follow-up", Col["Closed?"], "Yes = closed, Follow-up = pending")

# widths
widths = [11,20,14,14,22,11,13,10,20,11,13,13]
for i,w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 30

# ---------- Sheet 2: Summary ----------
s2 = wb.create_sheet("Summary")
s2["A1"] = "SiteCraft SA — Lead Tracker Summary"
s2["A1"].font = Font(bold=True, size=14, color=ACCENT)
s2.append([])
rows = [
    ("Total leads tracked",        f"=COUNTA(Leads!A2:A{N+1})"),
    ("Outreach sent",              f"=COUNTIF(Leads!{Col['Outreach sent?']}2:{Col['Outreach sent?']}{N+1},\"Yes\")"),
    ("Replies",                    f"=COUNTIF(Leads!{Col['Replied?']}2:{Col['Replied?']}{N+1},\"Yes\")"),
    ("Closed (won)",               f"=COUNTIF(Leads!{CLOSED}2:{CLOSED}{N+1},\"Yes\")"),
    ("Pending follow-ups",         f"=COUNTIF(Leads!{CLOSED}2:{CLOSED}{N+1},\"Follow-up\")"),
    ("Reply rate",                 f"=IF(B4=0,0,B5/B4)"),
    ("Close rate (of replies)",    f"=IF(B5=0,0,B6/B5)"),
    ("Total setup revenue (R)",    f"=SUMIF(Leads!{CLOSED}2:{CLOSED}{N+1},\"Yes\",Leads!{SETUP}2:{SETUP}{N+1})"),
    ("Monthly recurring (R/mo)",    f"=SUMIF(Leads!{CLOSED}2:{CLOSED}{N+1},\"Yes\",Leads!{MON}2:{MON}{N+1})"),
    ("Annual recurring (R/yr)",     "=B10*12"),
]
for label, formula in rows:
    s2.append([label, formula])
# style summary
for r in range(3, 3+len(rows)):
    s2.cell(row=r, column=1).font = Font(bold=True, size=10)
    s2.cell(row=r, column=1).fill = card_fill
    s2.cell(row=r, column=1).border = border
    s2.cell(row=r, column=2).border = border
    s2.cell(row=r, column=2).alignment = Alignment(horizontal="left")
    if "rate" in str(s2.cell(row=r,column=1).value).lower():
        s2.cell(row=r, column=2).number_format = "0.0%"
    else:
        s2.cell(row=r, column=2).number_format = "#,##0"
s2.column_dimensions["A"].width = 26
s2.column_dimensions["B"].width = 18

# ---------- Sheet 3: How to use ----------
s3 = wb.create_sheet("How to use")
s3["A1"] = "How to use this tracker"
s3["A1"].font = Font(bold=True, size=14, color=ACCENT)
tips = [
    "1. Prospect on Google Maps (maps.google.com): search '<trade> in <town>'.",
    "2. For each weak profile, add a row: business, trade, town, contact (WA link or phone).",
    "3. Red flags (col F): count how many of these apply — claim status, no site, <20 reviews, <5 photos, missing hours/phone, no posts, weak description, low rank.",
    "4. Send the WhatsApp opener, then set 'Outreach sent?' = Yes.",
    "5. When they reply, set 'Replied?' = Yes and send the free mini-audit.",
    "6. On a win, set 'Closed?' = Yes and fill Setup value (R1,500) + Monthly (R450). Pending = 'Follow-up'.",
    "7. The Summary tab auto-calculates reply rate, close rate, and revenue.",
    "8. Target: 20-30 prospects/week -> ~3-5 closes -> R4,500-R7,500 setup + R1,350-R2,250/mo.",
    "",
    "Outreach opener (WhatsApp, Click-to-Chat): https://wa.me/27XXXXXXXXX?text=...",
    "Free Skillshop course: search 'Google Business Profile course Skillshop'.",
    "Contact: WhatsApp +27 74 508 6001 · lehauthabang@gmail.com",
]
for t in tips:
    s3.append([t])
s3.column_dimensions["A"].width = 110
for r in range(2, 2+len(tips)):
    s3.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")

wb.save("sitecraft_sa_lead_tracker.xlsx")
print("XLSX built: sitecraft_sa_lead_tracker.xlsx")
